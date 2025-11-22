# app.py
# pip install flask openpyxl

from flask import Flask, request, send_file, jsonify
from openpyxl import load_workbook, Workbook
import os
import io
import datetime as dt

app = Flask(__name__)

# 업로드된 BOM/좌표 엑셀들이 들어있는 폴더
# 실제 환경에 맞게 수정하세요.
UPLOAD_DIR = r"./uploads"

# ---------- 공통 유틸 ----------

def sheet_to_aoa(ws):
    """openpyxl 워크시트를 2차원 배열(AoA)로 변환"""
    max_row = ws.max_row
    max_col = ws.max_column
    aoa = []
    for r in range(1, max_row + 1):
        row = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            row.append("" if v is None else v)
        aoa.append(row)
    return aoa

def find_header_row(aoa, headers):
    """
    headers: {키: 헤더텍스트}
    aoa에서 해당 헤더들을 모두 포함하는 행을 찾고
    각 키가 몇 번째 컬럼인지 colMap을 돌려준다.
    """
    for r, row in enumerate(aoa):
        name_to_col = {}
        for c, v in enumerate(row):
            if isinstance(v, str):
                name_to_col[v.strip()] = c

        ok = True
        for key, header_text in headers.items():
            if header_text not in name_to_col:
                ok = False
                break
        if ok:
            col_map = {key: name_to_col[h] for key, h in headers.items()}
            return {"headerRow": r, "colMap": col_map}
    return None

# ---------- BOM 파싱 ----------

def parse_bom_workbook(wb):
    want_headers = {
        "spec": "Specification(제품 용량값)",
        "pkg": "Package(제품 사이즈)",
        "ref": "Reference(제품 위치값)",
    }

    target_sheet = None
    aoa = None
    header_info = None

    for name in wb.sheetnames:
        ws = wb[name]
        arr = sheet_to_aoa(ws)
        info = find_header_row(arr, want_headers)
        if info:
            target_sheet = ws
            aoa = arr
            header_info = info
            break

    if not target_sheet or not header_info:
        raise ValueError(
            "BOM 시트에서 [Specification(제품 용량값) / Package(제품 사이즈) / Reference(제품 위치값)] 헤더를 찾지 못했습니다."
        )

    header_row = header_info["headerRow"]
    col_map = header_info["colMap"]

    bom_rows = []
    summary_by_spec = {}
    all_refs_set = set()

    for r in range(header_row + 1, len(aoa)):
        row = aoa[r]
        spec = str(row[col_map["spec"]]).strip() if col_map["spec"] < len(row) else ""
        pkg = str(row[col_map["pkg"]]).strip() if col_map["pkg"] < len(row) else ""
        ref_str = str(row[col_map["ref"]]).strip() if col_map["ref"] < len(row) else ""

        if not (spec or pkg or ref_str):
            continue

        refs = [s.strip() for s in ref_str.split(",") if s.strip()]
        qty = len(refs)

        pkg_spec = f"{pkg}-{spec}" if (pkg and spec) else (pkg or spec or "")

        bom_rows.append({
            "spec": spec,
            "pkg": pkg,
            "refs": refs,
            "pkgSpec": pkg_spec,
            "qty": qty,
        })

        if spec:
            summary_by_spec[spec] = summary_by_spec.get(spec, 0) + qty

        for rf in refs:
            all_refs_set.add(rf)

    summary_list = [{"spec": spec, "qty": qty} for spec, qty in summary_by_spec.items()]

    return {
        "bomRows": bom_rows,
        "summaryList": summary_list,
        "allRefsSet": all_refs_set,
    }

# ---------- 좌표 파싱 ----------

def parse_coord_workbook(wb):
    want_headers = {
        "ref": "Reference(제품 위치값)",
        "x": "좌표값 X축",
        "y": "좌표값 Y축",
        "rot": "좌표값 회전",
        "side": "Top/Bot",
    }

    target_sheet = None
    aoa = None
    header_info = None

    for name in wb.sheetnames:
        ws = wb[name]
        arr = sheet_to_aoa(ws)
        info = find_header_row(arr, want_headers)
        if info:
            target_sheet = ws
            aoa = arr
            header_info = info
            break

    if not target_sheet or not header_info:
        raise ValueError(
            "좌표 시트에서 [Reference(제품 위치값) / 좌표값 X축 / 좌표값 Y축 / 좌표값 회전 / Top/Bot] 헤더를 찾지 못했습니다."
        )

    header_row = header_info["headerRow"]
    col_map = header_info["colMap"]

    coord_map = {}  # key: ref, value: {x, y, rot, side}

    for r in range(header_row + 1, len(aoa)):
        row = aoa[r]
        ref = str(row[col_map["ref"]]).strip() if col_map["ref"] < len(row) else ""
        if not ref:
            continue

        x = row[col_map["x"]] if col_map["x"] < len(row) else ""
        y = row[col_map["y"]] if col_map["y"] < len(row) else ""
        rot = row[col_map["rot"]] if col_map["rot"] < len(row) else ""
        side_raw = row[col_map["side"]] if col_map["side"] < len(row) else ""
        side = str(side_raw).strip()

        if ref not in coord_map:
            coord_map[ref] = {
                "x": x,
                "y": y,
                "rot": rot,
                "side": side,
            }

    return coord_map

def merge_coord_maps(maps):
    merged = {}
    for m in maps:
        for ref, data in m.items():
            if ref not in merged:
                merged[ref] = data
    return merged

# ---------- 결과 엑셀 생성 로직 ----------

def build_result_workbook(bom_parsed, coord_map, base_name):
    wb_out = Workbook()
    ws_result = wb_out.active
    ws_result.title = "결과값 추출"

    # 1) 제품 목록표 영역 (A1~)
    summary_rows = []
    summary_rows.append(["제품 목록표", "", ""])
    summary_rows.append(["Specification(제품 용량값)", "수량", "비고"])
    for item in bom_parsed["summaryList"]:
        summary_rows.append([item["spec"], item["qty"], ""])

    for r_idx, row in enumerate(summary_rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws_result.cell(row=r_idx, column=c_idx, value=val)

    # 2) 결과값 테이블 (F4~)
    result_header = [
        "No.",
        "Reference(제품 위치값)",
        "좌표값 X축",
        "좌표값 Y축",
        "좌표값 회전",
        "Package-Specification",
        "Top/Bot",
    ]

    offset_row = 4
    offset_col = 6  # F = 6

    # 헤더
    for c_idx, val in enumerate(result_header, start=offset_col):
        ws_result.cell(row=offset_row, column=c_idx, value=val)

    # 데이터
    no = 1
    cur_row = offset_row + 1
    for b in bom_parsed["bomRows"]:
        for ref in b["refs"]:
            coord = coord_map.get(ref, {})
            ws_result.cell(row=cur_row, column=offset_col + 0, value=no)
            ws_result.cell(row=cur_row, column=offset_col + 1, value=ref)
            ws_result.cell(row=cur_row, column=offset_col + 2, value=coord.get("x", ""))
            ws_result.cell(row=cur_row, column=offset_col + 3, value=coord.get("y", ""))
            ws_result.cell(row=cur_row, column=offset_col + 4, value=coord.get("rot", ""))
            ws_result.cell(row=cur_row, column=offset_col + 5, value=b["pkgSpec"])
            ws_result.cell(row=cur_row, column=offset_col + 6, value=coord.get("side", ""))

            no += 1
            cur_row += 1

    # 3) 미삽 추출 시트
    ws_miss = wb_out.create_sheet("미삽 추출")
    miss_header = [
        "Reference(제품 위치값)",
        "좌표값 X축",
        "좌표값 Y축",
        "좌표값 회전",
        "Top/Bot",
    ]
    for c_idx, val in enumerate(miss_header, start=1):
        ws_miss.cell(row=1, column=c_idx, value=val)

    miss_row = 2
    for ref, coord in coord_map.items():
        if ref not in bom_parsed["allRefsSet"]:
            ws_miss.cell(row=miss_row, column=1, value=ref)
            ws_miss.cell(row=miss_row, column=2, value=coord.get("x", ""))
            ws_miss.cell(row=miss_row, column=3, value=coord.get("y", ""))
            ws_miss.cell(row=miss_row, column=4, value=coord.get("rot", ""))
            ws_miss.cell(row=miss_row, column=5, value=coord.get("side", ""))
            miss_row += 1

    return wb_out

# ---------- API 엔드포인트 ----------

@app.route("/api/smt/extract", methods=["POST"])
def api_smt_extract():
    """
    요청 JSON:
    {
      "bomFile": "저장된_BOM파일명.xlsx",
      "coordFiles": ["저장된_TOP좌표.xlsx", "저장된_BOT좌표.xlsx"]
    }
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({"ok": False, "message": "JSON 바디가 없습니다."}), 400

        bom_name = data.get("bomFile")
        coord_names = data.get("coordFiles") or []

        if not bom_name:
            return jsonify({"ok": False, "message": "bomFile 이 필요합니다."}), 400
        if not coord_names:
            return jsonify({"ok": False, "message": "coordFiles 가 비어 있습니다."}), 400

        bom_path = os.path.join(UPLOAD_DIR, bom_name)
        if not os.path.exists(bom_path):
            return jsonify({"ok": False, "message": f"BOM 파일을 찾을 수 없습니다: {bom_name}"}), 404

        coord_paths = []
        for name in coord_names:
            p = os.path.join(UPLOAD_DIR, name)
            if not os.path.exists(p):
                return jsonify({"ok": False, "message": f"좌표 파일을 찾을 수 없습니다: {name}"}), 404
            coord_paths.append(p)

        # 1) BOM 파싱
        bom_wb = load_workbook(bom_path, data_only=True)
        bom_parsed = parse_bom_workbook(bom_wb)

        # 2) 좌표 파싱 (여러 개 병합)
        coord_maps = []
        for p in coord_paths:
            wb = load_workbook(p, data_only=True)
            coord_maps.append(parse_coord_workbook(wb))
        coord_map = merge_coord_maps(coord_maps)

        # 3) 결과 워크북 생성
        base_name = os.path.splitext(os.path.basename(bom_name))[0]
        wb_out = build_result_workbook(bom_parsed, coord_map, base_name)

        # 4) 메모리로 저장 후 전송
        stream = io.BytesIO()
        wb_out.save(stream)
        stream.seek(0)

        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        download_name = f"{base_name}_결과_미삽_{ts}.xlsx"

        return send_file(
            stream,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        # 디버그용 로그
        print("ERROR in /api/smt/extract:", e)
        return jsonify({"ok": False, "message": str(e)}), 500


if __name__ == "__main__":
    # 개발용 실행
    app.run(host="0.0.0.0", port=5000, debug=True)
